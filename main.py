#csv extraido de: https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-abertos/arquivos/vdpb/vendas-derivados-petroleo-e-etanol/vendas-derivados-petroleo-etanol-m3-1990-2021.csv
#csv extraido de:https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-abertos/arquivos/vdpb/vct/vendas-oleo-diesel-tipo-m3-2013-2021.csv

import csv
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

from airflow import DAG
from datetime import datetime
from airflow.operators.python import PythonOperator
from airflow.operators.empty import EmptyOperator

csv_diesel_path = '/home/patricia/airflow/external_files/vendas-oleo-diesel-tipo-m3-2013-2023.csv'
csv_etanol_path = '/home/patricia/airflow/external_files/vendas-derivados-petroleo-etanol-m3-1990-2023.csv'
raw_file = '/home/patricia/airflow/external_files/vendas-combustiveis-m3.xls'
output_diesel_file = '/home/patricia/airflow/external_files/sales_diesel_uf_type.csv'
output_derivative_file = '/home/patricia/airflow/external_files/sales_derivative_fuels.csv'
raw_file_xls = '/home/patricia/airflow/external_files/vendas-combustiveis-m3.xls'
raw_file_xlsx = '/home/patricia/airflow/external_files/vendas-combustiveis-m3.xlsx'
schema = ['year_month', 'uf', 'product','unit','volume','created_at'] 
now = datetime.now()

def read_csv(csv_file, output_file):
    with open(csv_file) as f:
        reader_file = csv.reader(f, delimiter =";")
        ignore_old_header(reader_file)
        create_new_csv(reader_file, output_file)

def create_new_csv(reader_file, output_file):
    new_csv = csv.writer(open(output_file, 'w'), delimiter=';',quotechar='\'')
    write_new_header(new_csv)  
    for row in reader_file:
        if int(row[0]) <= 2020:
            new_row = [';'.join([row[0]+row[1], row[3],row[4],'m3',row[5], str(now)])] 
            new_csv.writerow(new_row)
        
def get_total_derivative_fuels_from_generated_csv():
    with open(output_derivative_file) as fin:
        next(fin, None)
        total = 0
        for row in csv.reader(fin, delimiter=";"):
            total += int(float(row[4].replace(',', '.')))
        return total

def get_total_diesel_from_generated_csv():
    with open(output_diesel_file) as fin:
        next(fin, None)
        total = 0
        for row in csv.reader(fin, delimiter=";"):
            total += int(float(row[4].replace(',', '.')))
        return total

def read_raw_xls_convert_xlsx(): 
    raw_data = pd.read_excel(raw_file_xls)
    raw_data.to_excel(raw_file_xlsx)

def get_total_derivative_fuels_from_raw_file():
    read_sheet = load_workbook(raw_file_xlsx)
    active_workbook = read_sheet.worksheets[0]
    total = 0
    for col in range(4, 25):
        cell_data = active_workbook.cell(row=66, column=col).value
        total += cell_data
    return total

def get_total_diesel_from_raw_file():
    read_sheet = load_workbook(raw_file_xlsx)
    active_workbook = read_sheet.worksheets[0]
    total = 0
    for col in range(4, 12):
        cell_data = active_workbook.cell(row=146, column=col).value
        total += cell_data
    return total

def read_csv_diesel_uf_type():
    read_csv(csv_diesel_path, output_diesel_file)

def read_csv_derivative_fuels():
    read_csv(csv_etanol_path, output_derivative_file)

def write_new_header(writer):
    writer.writerow([g for g in schema]) 

def ignore_old_header(reader_file):
    next(reader_file, None) #pula a primeira linha do csv, com o cabeçalho antigo

def compare_diesel_raw_file_generated_file(ti):
  total_raw = ti.xcom_pull(task_ids='get_total_diesel_from_raw_file')
  total_generated = ti.xcom_pull(task_ids='get_total_diesel_from_generated_csv')

  if (total_raw==total_generated):
    return 'Total de registros igual'
  return 'Total de registros diferente'

def compare_derivative_raw_file_generated_file(ti):
  total_raw = ti.xcom_pull(task_ids='get_total_derivative_fuels_from_raw_file')
  total_generated = ti.xcom_pull(task_ids='get_total_derivative_fuels_from_generated_csv')

  if (total_raw==total_generated):
    return 'Total de registros igual'
  return 'Total de registros diferente'


with DAG('test_read_xlsx', start_date=datetime(2023,10,9),
         schedule = '30 * * * *', catchup=False) as dag:

    start = EmptyOperator(
        task_id='start'
    )
    read_csv_diesel_uf_type = PythonOperator(
        task_id = 'read_csv_diesel_uf_type',
        python_callable = read_csv_diesel_uf_type,
    )
    read_csv_derivative_fuels = PythonOperator(
        task_id = 'read_csv_derivative_fuels',
        python_callable = read_csv_derivative_fuels,
    )
    get_total_derivative_fuels_from_generated_csv = PythonOperator(
        task_id = 'get_total_derivative_fuels_from_generated_csv',
        python_callable = get_total_derivative_fuels_from_generated_csv,
    )
    get_total_diesel_from_generated_csv = PythonOperator(
        task_id = 'get_total_diesel_from_generated_csv',
        python_callable = get_total_diesel_from_generated_csv,
    )
    read_raw_xls_convert_xlsx = PythonOperator(
        task_id = 'read_raw_xls_convert_xlsx',
        python_callable = read_raw_xls_convert_xlsx,
    )
    get_total_derivative_fuels_from_raw_file = PythonOperator(
        task_id = 'get_total_derivative_fuels_from_raw_file',
        python_callable = get_total_derivative_fuels_from_raw_file,
    )
    get_total_diesel_from_raw_file = PythonOperator(
        task_id = 'get_total_diesel_from_raw_file',
        python_callable = get_total_diesel_from_raw_file,
    )
    compare_diesel_raw_file_generated_file = PythonOperator(
        task_id = 'compare_diesel_raw_file_generated_file',
        python_callable = compare_diesel_raw_file_generated_file,
    )
    compare_derivative_raw_file_generated_file = PythonOperator(
        task_id = 'compare_derivative_raw_file_generated_file',
        python_callable = compare_derivative_raw_file_generated_file,
    )
    
    start >>  read_csv_derivative_fuels >> get_total_derivative_fuels_from_generated_csv 
    read_csv_diesel_uf_type >> get_total_diesel_from_generated_csv
    read_raw_xls_convert_xlsx >> get_total_derivative_fuels_from_raw_file  
    read_raw_xls_convert_xlsx >> get_total_diesel_from_raw_file
    get_total_diesel_from_raw_file >> compare_diesel_raw_file_generated_file
    get_total_derivative_fuels_from_raw_file >> compare_derivative_raw_file_generated_file